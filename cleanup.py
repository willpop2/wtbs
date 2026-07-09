"""
Step 2 of the WTBS pipeline: turn a raw transcript into a readable one with Claude.

Fixes proper nouns against glossary.xlsx, edits for readability (cuts filler and
backchannel, condenses, light rewrite for flow), italicizes work titles, and maps
speaker labels to real names. Processes the transcript in chunks for long episodes.

Usage:
    python cleanup.py transcripts/raw/my-episode.txt
    python cleanup.py transcripts/raw/my-episode.txt --bio bios/my-episode.txt --names A=Will B=Jimmy
    python cleanup.py transcripts/raw/my-episode.txt --bio bio.txt --model claude-opus-4-8

Output:
    transcripts/clean/<name>.txt
"""

import argparse
import os
import re
import sys
from pathlib import Path

import anthropic
from dotenv import load_dotenv
from openpyxl import load_workbook

ROOT = Path(__file__).parent
CLEAN_DIR = ROOT / "transcripts" / "clean"

# Approx words per chunk sent to Claude. Small enough to keep corrections
# focused and stay well under output limits; large enough to be efficient.
CHUNK_WORDS = 2500

# Token-usage accumulator across all requests; reported at the end so cache hits
# are visible (cache_read > 0 means prompt caching is working).
USAGE = {"input": 0, "cache_write": 0, "cache_read": 0, "output": 0}

SYSTEM_TEMPLATE = """You are an editor preparing a spoken interview for publication as a readable \
transcript on the WTBS website. WTBS ("Waiting to Be Signed") features long-form interviews with \
generative, on-chain, and NFT artists. You turn raw speech-to-text output into a transcript that is \
pleasant and clear to READ, while staying true to what each speaker actually said.

Your goal: a polished, engaging reading experience. Edit ASSERTIVELY -- prioritize clarity, flow, \
and concision over staying literal to the spoken words. A reader should never be confused, annoyed, \
or bored.

DO:
- Fix misspelled or misheard proper nouns: artist names, handles, platforms, blockchains, studios, \
galleries, tools, and technical terms. Use the glossary below as the authoritative spelling \
(including phonetic spellings, e.g. "ciphered" -> "Ciphrd").
- Italicize the titles of specific creative works -- individual artworks, generative art projects, \
collections, and games -- using Markdown asterisks, e.g. *Proscenium*, *Architectonica*, *vibes*, \
*the reliquary*, *Terraforms*, *Fidenza*, *EverWing*. Do NOT italicize platforms, marketplaces, \
blockchains, studios, galleries, companies, people, or handles -- Art Blocks, Tezos, Ethereum, \
OpenSea, fx(hash), Verse, and similar stay in plain text.
- Aggressively cut filler and verbal tics ("um", "uh", "like", "you know", "I mean", "kind of", \
"sort of", "right?", "I guess", "basically") wherever they don't carry meaning, plus false starts, \
stutters, self-corrections, and hedging.
- Remove standalone backchannel / acknowledgment turns -- a speaker turn whose whole content is a \
brief verbal nod with no substance ("Yeah.", "Oh, yeah.", "Mhm.", "Hmm.", "Right.", "Sure.", \
"Exactly.", "Totally.", "For sure.", "Nice."). These work in audio but fragment the other person's \
answer in text. When removing such an interjection leaves two adjacent turns by the same speaker, \
merge them into one continuous turn (one label). Keep a short affirmation only when it is the actual \
substantive answer to a question.
- Condense rambling, circular, or repetitive passages to their clear point. If a speaker makes the \
same point three times, make it once, well.
- Freely restructure: reorder clauses, merge or split sentences, and rewrite meandering spoken \
syntax into clean, flowing written prose. Fix all grammar and punctuation. Break long turns into \
paragraphs.
- Aim for prose a reader would enjoy in a well-edited published interview.

PRESERVE (do not cross these lines):
- Every fact, opinion, anecdote, argument, and conclusion the speaker actually expressed. Never add \
information they did not say, never change their meaning or stance, and never drop a substantive \
point -- even while condensing how it is said.
- The speaker's authentic voice and personality: keep their humor, memorable turns of phrase, and \
conversational warmth. Polished, not sterilized -- it should read like the best version of how they \
talk, not a corporate press release.
- First person, and the exact speaker turns. Keep every speaker label exactly as given. Never merge, \
drop, or reattribute a turn between different speakers.

When condensing, protect substance and voice above literal wording.

Return ONLY the edited transcript. No preamble, no commentary, no code fences.

GLOSSARY (authoritative spellings):
{glossary}"""


def load_keyterms(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    terms = [str(c).strip() for (c,) in ws.iter_rows(min_col=1, max_col=1, values_only=True)
             if c is not None and str(c).strip()]
    wb.close()
    return terms


def chunk_blocks(blocks: list[str], max_words: int) -> list[str]:
    """Group speaker blocks into chunks under a word budget (never split a block)."""
    chunks, current, count = [], [], 0
    for block in blocks:
        words = len(block.split())
        if current and count + words > max_words:
            chunks.append("\n\n".join(current))
            current, count = [], 0
        current.append(block)
        count += words
    if current:
        chunks.append("\n\n".join(current))
    return chunks


def merge_same_speaker(text: str, labels: set) -> str:
    """Merge consecutive turns by the same speaker into one (safety net for
    removed interjections). Continuation paragraphs (no label) are left as-is."""
    ordered = sorted(labels, key=len, reverse=True)
    pat = re.compile(r"^(" + "|".join(re.escape(l) for l in ordered) + r"):\s*(.*)$",
                     re.DOTALL)
    out, last = [], None
    for b in (b for b in text.split("\n\n") if b.strip()):
        m = pat.match(b.strip())
        if m and m.group(1) == last and out:
            out[-1] = out[-1].rstrip() + " " + m.group(2).strip()
        else:
            out.append(b.strip())
            if m:
                last = m.group(1)
    return "\n\n".join(out)


BASE_INSTRUCTION = ("Edit this transcript excerpt per your rules and return only the "
                    "edited text:\n\n")
# Used to retry a chunk the model over-condensed: emphasize preservation.
PRESERVE_INSTRUCTION = (
    "Edit this transcript excerpt and return only the edited text. IMPORTANT: a previous "
    "attempt cut far too much. This passage is SUBSTANTIVE, not repetitive filler -- keep "
    "every point, detail, argument, and example the speakers make. Only remove filler words "
    "and verbal tics, fix names/terms, italicize work titles, and tidy grammar. Keep the "
    "result close to the original length.\n\n")


def edit_chunk(client, model: str, system: str, chunk: str, depth: int = 0):
    """Edit one chunk. Retry with a preservation nudge if the model collapses or
    truncates it; if it still collapses, split into smaller pieces and edit each
    (the model can't compress a whole span to its gist if it only sees a piece).
    Returns (text, out_words, ok)."""
    in_w = len(chunk.split())
    best = ""
    for instruction in (BASE_INSTRUCTION, PRESERVE_INSTRUCTION):
        msg = client.messages.create(
            model=model, max_tokens=16000, thinking={"type": "disabled"},
            system=system, messages=[{"role": "user", "content": instruction + chunk}],
        )
        u = msg.usage
        USAGE["input"] += u.input_tokens
        USAGE["cache_write"] += getattr(u, "cache_creation_input_tokens", 0) or 0
        USAGE["cache_read"] += getattr(u, "cache_read_input_tokens", 0) or 0
        USAGE["output"] += u.output_tokens
        txt = "".join(b.text for b in msg.content if b.type == "text").strip()
        out_w = len(txt.split())
        if len(txt.split()) > len(best.split()):
            best = txt
        ok = bool(txt) and msg.stop_reason != "max_tokens" and not (
            in_w >= 400 and out_w < 0.35 * in_w)
        if ok:
            return txt, out_w, True

    # Still collapsing: divide and conquer down to smaller spans.
    blocks = [b for b in chunk.split("\n\n") if b.strip()]
    if len(blocks) > 1 and depth < 5:
        mid = len(blocks) // 2
        lt, _, lok = edit_chunk(client, model, system, "\n\n".join(blocks[:mid]), depth + 1)
        rt, _, rok = edit_chunk(client, model, system, "\n\n".join(blocks[mid:]), depth + 1)
        combined = (lt + "\n\n" + rt).strip()
        return combined, len(combined.split()), (lok and rok)
    return best, len(best.split()), False


def main() -> None:
    ap = argparse.ArgumentParser(description="Correct a transcript with Claude.")
    ap.add_argument("transcript", help="Path to a raw transcript .txt file.")
    ap.add_argument("--bio", default=None,
                    help="Guest bio: inline text, or a path to a .txt file.")
    ap.add_argument("--model", default="claude-sonnet-5",
                    help="Anthropic model id (default: claude-sonnet-5).")
    ap.add_argument("--names", nargs="*", default=[], metavar="LABEL=NAME",
                    help='Rename speaker labels, e.g. --names A=Will B=Jimmy')
    ap.add_argument("--out-dir", default=None,
                    help="Output directory (default: transcripts/clean).")
    args = ap.parse_args()

    load_dotenv(ROOT / ".env")
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        sys.exit("ANTHROPIC_API_KEY missing — check your .env file.")

    src = Path(args.transcript)
    if not src.exists():
        sys.exit(f"Transcript not found: {src}")

    bio_text = ""
    if args.bio:
        bio_path = Path(args.bio)
        raw_bio = bio_path.read_text(encoding="utf-8") if bio_path.exists() else args.bio
        bio_text = f"\n\nGUEST / EPISODE CONTEXT:\n{raw_bio.strip()}"

    keyterms = load_keyterms(ROOT / "glossary.xlsx")
    glossary_str = "\n".join(f"- {t}" for t in keyterms)
    # The big instructions+glossary block is identical across every chunk and
    # every episode, so cache it (read at ~0.1x after the first write). The
    # per-episode bio is appended as a separate, uncached block after it.
    system = [{"type": "text",
               "text": SYSTEM_TEMPLATE.format(glossary=glossary_str),
               "cache_control": {"type": "ephemeral"}}]
    if bio_text:
        system.append({"type": "text", "text": bio_text.strip()})

    # Build "Speaker A" -> "Will" style mapping and apply to block labels.
    name_map = {}
    for pair in args.names:
        if "=" not in pair:
            continue
        label, name = (p.strip() for p in pair.split("=", 1))
        key = label if label.lower().startswith("speaker") else f"Speaker {label}"
        name_map[key] = name

    def rename(block: str) -> str:
        for label, name in name_map.items():
            if block.startswith(label + ":"):
                return name + ":" + block[len(label) + 1:]
        return block

    text = src.read_text(encoding="utf-8")
    blocks = [rename(b) for b in text.split("\n\n") if b.strip()]
    if name_map:
        print("Speaker labels: " + ", ".join(f"{k} -> {v}" for k, v in name_map.items()))
    chunks = chunk_blocks(blocks, CHUNK_WORDS)
    print(f"Loaded {len(keyterms)} keyterms. Correcting {len(blocks)} blocks "
          f"in {len(chunks)} chunk(s) with {args.model} ...")

    client = anthropic.Anthropic(api_key=api_key)
    corrected_parts = []
    for i, chunk in enumerate(chunks, 1):
        text_out, out_w, ok = edit_chunk(client, args.model, system, chunk)
        in_w = len(chunk.split())
        flag = "" if ok else "  <-- COLLAPSED after retries; keeping best attempt"
        print(f"  chunk {i}/{len(chunks)}: {in_w} -> {out_w} words{flag}")
        # Fail loud only if even the preservation retries produced nothing usable.
        if not text_out:
            sys.exit(f"chunk {i} produced no output; aborting to avoid content loss.")
        corrected_parts.append(text_out)

    final_text = "\n\n".join(corrected_parts)
    labels = set(name_map.values()) | {f"Speaker {c}" for c in "ABCDEFGH"}
    final_text = merge_same_speaker(final_text, labels)

    out_dir = Path(args.out_dir) if args.out_dir else CLEAN_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / src.name
    out.write_text(final_text, encoding="utf-8")
    print(f"\nDone. Cleaned transcript: {out}")
    print(f"Tokens — input: {USAGE['input']:,} · cache write: {USAGE['cache_write']:,} "
          f"· cache read: {USAGE['cache_read']:,} · output: {USAGE['output']:,}")


if __name__ == "__main__":
    main()
