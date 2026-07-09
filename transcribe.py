"""
Step 1 of the WTBS pipeline: transcribe an episode with AssemblyAI.

Uses the slam-1 speech model with keyterms (from glossary.xlsx) + speaker
diarization, tuned for proper-noun-heavy artist interviews.

Usage:
    python transcribe.py audio/my-episode.mp3
    python transcribe.py audio/my-episode.mp3 --speakers 2

Outputs (named after the audio file):
    transcripts/raw/<name>.json   full AssemblyAI response (kept for reference)
    transcripts/raw/<name>.txt    readable, speaker-labeled transcript
"""

import argparse
import json
import sys
from pathlib import Path

import assemblyai as aai
from dotenv import load_dotenv
from openpyxl import load_workbook

ROOT = Path(__file__).parent
RAW_DIR = ROOT / "transcripts" / "raw"


def load_keyterms(path: Path) -> list[str]:
    """Read the glossary: one term per row in column A."""
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    terms = []
    for (cell,) in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        if cell is None:
            continue
        term = str(cell).strip()
        if term:
            terms.append(term)
    wb.close()
    return terms


def main() -> None:
    ap = argparse.ArgumentParser(description="Transcribe an episode with AssemblyAI.")
    ap.add_argument("audio", help="Path to the audio file (mp3/wav).")
    ap.add_argument(
        "--speakers",
        type=int,
        default=None,
        help="Number of speakers you expect (optional; improves diarization).",
    )
    args = ap.parse_args()

    load_dotenv(ROOT / ".env")
    import os

    api_key = os.environ.get("ASSEMBLYAI_API_KEY")
    if not api_key:
        sys.exit("ASSEMBLYAI_API_KEY missing — check your .env file.")
    aai.settings.api_key = api_key

    audio_path = Path(args.audio)
    if not audio_path.exists():
        sys.exit(f"Audio file not found: {audio_path}")

    keyterms = load_keyterms(ROOT / "glossary.xlsx")
    print(f"Loaded {len(keyterms)} keyterms from glossary.xlsx")
    print(f"Transcribing {audio_path.name} with universal-3-5-pro + diarization ...")
    print("(This runs on AssemblyAI's servers; a 1-hour episode takes a few minutes.)")

    config = aai.TranscriptionConfig(
        speech_models=["universal-3-5-pro"],
        keyterms_prompt=keyterms,
        speaker_labels=True,
        speakers_expected=args.speakers,
        punctuate=True,
        format_text=True,
    )

    transcript = aai.Transcriber().transcribe(str(audio_path), config)

    if transcript.status == aai.TranscriptStatus.error:
        sys.exit(f"Transcription failed: {transcript.error}")

    RAW_DIR.mkdir(parents=True, exist_ok=True)
    stem = audio_path.stem

    # Full response for reference / re-processing.
    json_path = RAW_DIR / f"{stem}.json"
    json_path.write_text(json.dumps(transcript.json_response, indent=2), encoding="utf-8")

    # Readable, speaker-labeled transcript.
    txt_path = RAW_DIR / f"{stem}.txt"
    lines = []
    if transcript.utterances:
        for u in transcript.utterances:
            lines.append(f"Speaker {u.speaker}: {u.text}")
    else:
        lines.append(transcript.text or "")
    txt_path.write_text("\n\n".join(lines), encoding="utf-8")

    n_speakers = len({u.speaker for u in transcript.utterances}) if transcript.utterances else 0
    print("\nDone.")
    print(f"  Speakers detected: {n_speakers}")
    print(f"  Raw JSON : {json_path}")
    print(f"  Transcript: {txt_path}")
    print("\nNext: python cleanup.py transcripts/raw/%s.txt" % stem)


if __name__ == "__main__":
    main()
