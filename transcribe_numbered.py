"""
Raw-only transcription of the NUMBERED (non-interview) WTBS episodes.

Downloads + transcribes each numbered episode (AssemblyAI universal-3-5-pro,
keyterms + diarization) and saves ONLY the raw transcript -- no Claude cleanup,
since publishing these is still undecided. Kept separate from the interview
material. Resumable, and stops cleanly when AssemblyAI credit runs out.

Usage:
    python transcribe_numbered.py

Outputs:
    audio_numbered/<slug>.mp3
    transcripts/raw_numbered/<slug>.json  + <slug>.txt
"""

import os
import re
import sys
import urllib.request
import xml.etree.ElementTree as ET
from email.utils import parsedate_to_datetime
from pathlib import Path

import assemblyai as aai
from dotenv import load_dotenv
from openpyxl import load_workbook

ROOT = Path(__file__).parent
FEED = ROOT / "scratch_feed.xml"
AUDIO = ROOT / "audio_numbered"
RAW = ROOT / "transcripts" / "raw_numbered"
NS = {"i": "http://www.itunes.com/dtds/podcast-1.0.dtd"}
CREDIT_HINTS = ("credit", "balance", "insufficient", "fund", "payment", "quota", "limit")


def load_keyterms() -> list:
    wb = load_workbook(ROOT / "glossary.xlsx", read_only=True)
    ws = wb.active
    terms = [str(c).strip() for (c,) in ws.iter_rows(min_col=1, max_col=1, values_only=True)
             if c is not None and str(c).strip()]
    wb.close()
    return terms


def slugify(title: str, seen: set) -> str:
    m = re.match(r"^\s*(E\d+)\b", title)
    base = m.group(1) if m else re.sub(r"[^A-Za-z0-9]+", "_", title)[:40].strip("_") or "ep"
    slug, i = base, 2
    while slug.lower() in seen:
        slug = f"{base}_{i}"; i += 1
    seen.add(slug.lower())
    return slug


def numbered_episodes() -> list:
    root = ET.parse(FEED).getroot()
    eps = []
    for it in root.findall(".//item"):
        title = (it.findtext("title") or "").strip()
        if re.search(r"\binterview\b", title, re.I):
            continue  # skip interviews (handled separately)
        enc = it.find("enclosure")
        if enc is None:
            continue
        dt = parsedate_to_datetime(it.findtext("pubDate")) if it.findtext("pubDate") else None
        eps.append({"title": title, "url": enc.get("url"),
                    "sort": dt.timestamp() if dt else 0})
    eps.sort(key=lambda e: e["sort"], reverse=True)  # newest first
    seen = set()
    for e in eps:
        e["slug"] = slugify(e["title"], seen)
    return eps


def download(url: str, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    tmp = dest.with_suffix(dest.suffix + ".part")
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=120) as r, tmp.open("wb") as f:
        while chunk := r.read(1 << 20):
            f.write(chunk)
    tmp.replace(dest)


def main() -> None:
    load_dotenv(ROOT / ".env")
    key = os.environ.get("ASSEMBLYAI_API_KEY")
    if not key:
        sys.exit("ASSEMBLYAI_API_KEY missing")
    aai.settings.api_key = key

    keyterms = load_keyterms()
    config = aai.TranscriptionConfig(
        speech_models=["universal-3-5-pro"], keyterms_prompt=keyterms,
        speaker_labels=True, punctuate=True, format_text=True)
    transcriber = aai.Transcriber()

    eps = numbered_episodes()
    RAW.mkdir(parents=True, exist_ok=True)
    done = new = 0
    consecutive_fail = 0
    print(f"{len(eps)} numbered episodes found. Transcribing (raw only) newest-first.\n")

    for i, e in enumerate(eps, 1):
        slug = e["slug"]
        txt = RAW / f"{slug}.txt"
        if txt.exists():
            done += 1
            continue
        audio = AUDIO / f"{slug}.mp3"
        print(f"[{i}/{len(eps)}] {slug} — {e['title'][:55]}")
        try:
            if not audio.exists():
                print("  downloading ...")
                download(e["url"], audio)
            print("  transcribing ...")
            t = transcriber.transcribe(str(audio), config)
            if t.status == aai.TranscriptStatus.error:
                raise RuntimeError(t.error or "unknown error")
            (RAW / f"{slug}.json").write_text(
                __import__("json").dumps(t.json_response, indent=2), encoding="utf-8")
            lines = [f"Speaker {u.speaker}: {u.text}" for u in (t.utterances or [])] or [t.text or ""]
            txt.write_text("\n\n".join(lines), encoding="utf-8")
            new += 1
            consecutive_fail = 0
            print("  done.\n")
        except Exception as ex:
            msg = str(ex).lower()
            consecutive_fail += 1
            if any(h in msg for h in CREDIT_HINTS):
                print(f"  STOP — looks like AssemblyAI credit is exhausted: {ex}\n")
                break
            print(f"  ERROR: {ex}\n")
            if consecutive_fail >= 3:
                print("  STOP — 3 consecutive failures (likely credit exhausted).\n")
                break

    print(f"Finished. newly transcribed this run: {new}; already had: {done}; "
          f"total raw numbered: {len(list(RAW.glob('*.txt')))}/{len(eps)}")


if __name__ == "__main__":
    main()
